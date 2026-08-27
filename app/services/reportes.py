"""
=========================================================
SWAV
SERVICIO DE REPORTES HISTORICOS
=========================================================

Fuentes oficiales:
- HistoricoRegistro
- HistoricoPPU
- HistoricoExpedicion

Este servicio:
- NO lee R1.6
- NO usa Expedicion como fuente historica
- NO recalcula velocidades
- NO recalcula clasificaciones
=========================================================
"""

from io import BytesIO
from datetime import datetime, date

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from app.models import (
    HistoricoRegistro,
    HistoricoPPU,
    HistoricoExpedicion,
)


# =========================================================
# UTILIDADES
# =========================================================

def _normalizar_texto(valor):

    if valor is None:
        return ""

    return str(valor).strip()


def _normalizar_mayuscula(valor):

    return _normalizar_texto(
        valor
    ).upper()


def _convertir_fecha(valor):

    if valor in (
        None,
        ""
    ):
        return None

    if isinstance(
        valor,
        datetime
    ):
        return valor.date()

    if isinstance(
        valor,
        date
    ):
        return valor

    texto = str(
        valor
    ).strip()

    formatos = (
        "%Y-%m-%d",
        "%d/%m/%Y",
    )

    for formato in formatos:

        try:

            return datetime.strptime(
                texto,
                formato
            ).date()

        except ValueError:

            pass

    raise ValueError(
        f"Fecha invalida: {valor}"
    )


# =========================================================
# FILTROS HISTORICO REGISTRO
# =========================================================

def _aplicar_filtros_registro(
    consulta,
    fecha_desde=None,
    fecha_hasta=None,
    unidad=None,
    tipo_dia=None,
    servicio_usuario=None,
    servicio_empresa=None,
    indicador=None,
    clasificacion=None,
):

    fecha_desde = _convertir_fecha(
        fecha_desde
    )

    fecha_hasta = _convertir_fecha(
        fecha_hasta
    )

    if fecha_desde is not None:

        consulta = consulta.filter(
            HistoricoRegistro.fecha_operacional
            >= fecha_desde
        )

    if fecha_hasta is not None:

        consulta = consulta.filter(
            HistoricoRegistro.fecha_operacional
            <= fecha_hasta
        )

    if unidad:

        consulta = consulta.filter(
            HistoricoRegistro.unidad
            == _normalizar_mayuscula(
                unidad
            )
        )

    if tipo_dia:

        tipo_dia_normalizado = (
            _normalizar_mayuscula(
                tipo_dia
            )
        )

        # =============================================
        # NORMALIZACION OPERACIONAL TIPO DE DIA
        # =============================================
        # LABORAL agrupa registros historicos
        # almacenados como LABORAL o DIA NORMAL.
        # =============================================

        if tipo_dia_normalizado == "LABORAL":

            consulta = consulta.filter(
                HistoricoRegistro.tipo_dia.in_(
                    [
                        "LABORAL",
                        "DIA NORMAL",
                    ]
                )
            )

        else:

            consulta = consulta.filter(
                HistoricoRegistro.tipo_dia
                == tipo_dia_normalizado
            )

    if servicio_usuario:

        consulta = consulta.filter(
            HistoricoRegistro.servicio
            == _normalizar_texto(
                servicio_usuario
            )
        )

    if servicio_empresa:

        consulta = consulta.filter(
            HistoricoRegistro.ruta
            == _normalizar_texto(
                servicio_empresa
            )
        )

    if indicador:

        indicador_norm = (
            _normalizar_mayuscula(
                indicador
            )
        )

        if indicador_norm not in (
            "",
            "TODOS",
            "TODAS",
        ):

            consulta = consulta.filter(
                HistoricoRegistro
                .indicador_tiempo_espera
                == indicador_norm
            )

    if clasificacion:

        clasificacion_norm = (
            _normalizar_mayuscula(
                clasificacion
            )
        )

        if clasificacion_norm not in (
            "",
            "TODOS",
            "TODAS",
        ):

            consulta = consulta.filter(
                HistoricoRegistro.clasificacion
                == clasificacion_norm
            )

    return consulta


# =========================================================
# RESUMEN HISTORICO
# =========================================================

def consultar_resumen_historico(
    db,
    fecha_desde=None,
    fecha_hasta=None,
    unidad=None,
    tipo_dia=None,
    servicio_usuario=None,
    servicio_empresa=None,
    indicador=None,
    clasificacion=None,
):

    consulta = db.query(
        HistoricoRegistro
    )

    consulta = _aplicar_filtros_registro(
        consulta=consulta,
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta,
        unidad=unidad,
        tipo_dia=tipo_dia,
        servicio_usuario=servicio_usuario,
        servicio_empresa=servicio_empresa,
        indicador=indicador,
        clasificacion=clasificacion,
    )

    return (
        consulta
        .order_by(
            HistoricoRegistro.fecha_operacional,
            HistoricoRegistro.unidad,
            HistoricoRegistro.servicio,
            HistoricoRegistro.ruta,
            HistoricoRegistro.periodo,
        )
        .all()
    )


# =========================================================
# IDS HISTORICOS
# =========================================================

def _obtener_ids_historicos(
    db,
    fecha_desde=None,
    fecha_hasta=None,
    unidad=None,
    tipo_dia=None,
    servicio_usuario=None,
    servicio_empresa=None,
    indicador=None,
    clasificacion=None,
):

    registros = (
        consultar_resumen_historico(
            db=db,
            fecha_desde=fecha_desde,
            fecha_hasta=fecha_hasta,
            unidad=unidad,
            tipo_dia=tipo_dia,
            servicio_usuario=servicio_usuario,
            servicio_empresa=servicio_empresa,
            indicador=indicador,
            clasificacion=clasificacion,
        )
    )

    return [
        registro.id
        for registro in registros
    ]


# =========================================================
# DETALLE PPU
# =========================================================

def consultar_ppu_historico(
    db,
    fecha_desde=None,
    fecha_hasta=None,
    unidad=None,
    tipo_dia=None,
    servicio_usuario=None,
    servicio_empresa=None,
    indicador=None,
    clasificacion=None,
):

    ids = _obtener_ids_historicos(
        db=db,
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta,
        unidad=unidad,
        tipo_dia=tipo_dia,
        servicio_usuario=servicio_usuario,
        servicio_empresa=servicio_empresa,
        indicador=indicador,
        clasificacion=clasificacion,
    )

    if not ids:

        return []

    return (
        db.query(
            HistoricoPPU
        )
        .filter(
            HistoricoPPU.historico_id.in_(
                ids
            ),
            HistoricoPPU.velocidad_real.isnot(
                None
            ),
            HistoricoPPU.velocidad_real > 0,
        )
        .order_by(
            HistoricoPPU.fecha_operacional,
            HistoricoPPU.servicio,
            HistoricoPPU.ruta,
            HistoricoPPU.periodo,
            HistoricoPPU.inicio_servicio,
        )
        .all()
    )


# =========================================================
# EXPEDICIONES HISTORICAS
# =========================================================

def consultar_expediciones_historicas(
    db,
    fecha_desde=None,
    fecha_hasta=None,
    unidad=None,
    tipo_dia=None,
    servicio_usuario=None,
    servicio_empresa=None,
    indicador=None,
    clasificacion=None,
):

    consulta = db.query(
        HistoricoExpedicion
    )

    fecha_desde = _convertir_fecha(
        fecha_desde
    )

    fecha_hasta = _convertir_fecha(
        fecha_hasta
    )

    if fecha_desde is not None:

        consulta = consulta.filter(
            HistoricoExpedicion.fecha_operacional
            >= fecha_desde
        )

    if fecha_hasta is not None:

        consulta = consulta.filter(
            HistoricoExpedicion.fecha_operacional
            <= fecha_hasta
        )

    if unidad:

        consulta = consulta.filter(
            HistoricoExpedicion.unidad
            == _normalizar_mayuscula(
                unidad
            )
        )

    if tipo_dia:

        tipo_dia_normalizado = (
            _normalizar_mayuscula(
                tipo_dia
            )
        )

        # =============================================
        # NORMALIZACION OPERACIONAL TIPO DE DIA
        # =============================================

        if tipo_dia_normalizado == "LABORAL":

            consulta = consulta.filter(
                HistoricoExpedicion.tipo_dia.in_(
                    [
                        "LABORAL",
                        "DIA NORMAL",
                    ]
                )
            )

        else:

            consulta = consulta.filter(
                HistoricoExpedicion.tipo_dia
                == tipo_dia_normalizado
            )

    if servicio_usuario:

        consulta = consulta.filter(
            HistoricoExpedicion.servicio
            == _normalizar_texto(
                servicio_usuario
            )
        )

    if servicio_empresa:

        consulta = consulta.filter(
            HistoricoExpedicion.ruta
            == _normalizar_texto(
                servicio_empresa
            )
        )

    if indicador:

        indicador_norm = (
            _normalizar_mayuscula(
                indicador
            )
        )

        if indicador_norm not in (
            "",
            "TODOS",
            "TODAS",
        ):

            consulta = consulta.filter(
                HistoricoExpedicion
                .indicador_tiempo_espera
                == indicador_norm
            )

    if clasificacion:

        clasificacion_norm = (
            _normalizar_mayuscula(
                clasificacion
            )
        )

        if clasificacion_norm not in (
            "",
            "TODOS",
            "TODAS",
        ):

            consulta = consulta.filter(
                HistoricoExpedicion.clasificacion
                == clasificacion_norm
            )

    consulta = consulta.filter(
        HistoricoExpedicion.velocidad_real.isnot(
            None
        ),
        HistoricoExpedicion.velocidad_real > 0,
    )

    return (
        consulta
        .order_by(
            HistoricoExpedicion.fecha_operacional,
            HistoricoExpedicion.servicio,
            HistoricoExpedicion.ruta,
            HistoricoExpedicion.periodo,
            HistoricoExpedicion.inicio_servicio,
        )
        .all()
    )


# =========================================================
# FILTROS DISPONIBLES
# =========================================================

def obtener_filtros_reportes(
    db
):

    fechas = [
        fila[0]
        for fila in (
            db.query(
                HistoricoRegistro.fecha_operacional
            )
            .distinct()
            .order_by(
                HistoricoRegistro.fecha_operacional
            )
            .all()
        )
        if fila[0] is not None
    ]

    unidades = [
        fila[0]
        for fila in (
            db.query(
                HistoricoRegistro.unidad
            )
            .filter(
                HistoricoRegistro.unidad.isnot(
                    None
                )
            )
            .distinct()
            .order_by(
                HistoricoRegistro.unidad
            )
            .all()
        )
    ]

    # =============================================
    # TIPOS DE DIA OPERACIONALES
    # =============================================
    #
    # No exponer nombres tecnicos provenientes
    # directamente del historico.
    #
    # DIA NORMAL se considera LABORAL.
    # =============================================

    tipos_dia = [
        "LABORAL",
        "SABADO",
        "DOMINGO",
    ]

    servicios = [
        fila[0]
        for fila in (
            db.query(
                HistoricoRegistro.servicio
            )
            .filter(
                HistoricoRegistro.servicio.isnot(
                    None
                )
            )
            .distinct()
            .order_by(
                HistoricoRegistro.servicio
            )
            .all()
        )
    ]

    rutas = [
        fila[0]
        for fila in (
            db.query(
                HistoricoRegistro.ruta
            )
            .filter(
                HistoricoRegistro.ruta.isnot(
                    None
                )
            )
            .distinct()
            .order_by(
                HistoricoRegistro.ruta
            )
            .all()
        )
    ]

    return {

        "fechas": [
            fecha.isoformat()
            for fecha in fechas
        ],

        "unidades": unidades,

        "tipos_dia": tipos_dia,

        "servicios": servicios,

        "rutas": rutas,

        "indicadores": [
            "IP",
            "IE",
        ],

        "clasificaciones": [
            "OK",
            "SIMPLE",
            "COMPLEJO",
        ],
    }


# =========================================================
# EXCEL - UTILIDADES
# =========================================================

def _ajustar_columnas(
    ws
):

    for columna in ws.columns:

        maximo = 0

        letra = get_column_letter(
            columna[0].column
        )

        for celda in columna:

            if celda.value is None:

                continue

            largo = len(
                str(
                    celda.value
                )
            )

            if largo > maximo:

                maximo = largo

        ws.column_dimensions[
            letra
        ].width = min(
            maximo + 2,
            40
        )


def _formatear_cabecera(
    ws
):

    for celda in ws[1]:

        celda.font = Font(
            bold=True
        )

        celda.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    ws.freeze_panes = "A2"

    ws.auto_filter.ref = (
        ws.dimensions
    )


# =========================================================
# GENERAR EXCEL
# =========================================================

def generar_excel_historico(
    db,
    fecha_desde=None,
    fecha_hasta=None,
    unidad=None,
    tipo_dia=None,
    servicio_usuario=None,
    servicio_empresa=None,
    indicador=None,
    clasificacion=None,
):

    registros = consultar_resumen_historico(
        db=db,
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta,
        unidad=unidad,
        tipo_dia=tipo_dia,
        servicio_usuario=servicio_usuario,
        servicio_empresa=servicio_empresa,
        indicador=indicador,
        clasificacion=clasificacion,
    )

    ppus = consultar_ppu_historico(
        db=db,
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta,
        unidad=unidad,
        tipo_dia=tipo_dia,
        servicio_usuario=servicio_usuario,
        servicio_empresa=servicio_empresa,
        indicador=indicador,
        clasificacion=clasificacion,
    )

    wb = Workbook()

    # =====================================================
    # HOJA RESUMEN
    # =====================================================

    ws = wb.active
    ws.title = "Resumen"

    encabezados_resumen = [
        "ID",
        "Fecha",
        "Unidad",
        "Empresa",
        "Tipo Día",
        "Servicio Usuario",
        "Código TS",
        "Servicio Empresa",
        "Ruta Normalizada",
        "Sentido",
        "Período",
        "Expediciones",
        "Buses",
        "Velocidad Real",
        "Velocidad Teórica",
        "Reducción %",
        "Indicador",
        "Clasificación",
        "Estado",
        "Informar",
        "Observación",
        "Archivo Origen",
    ]

    ws.append(
        encabezados_resumen
    )

    for r in registros:

        ws.append(
            [
                r.id,

                (
                    r.fecha_operacional.isoformat()
                    if r.fecha_operacional
                    else None
                ),

                r.unidad,
                r.empresa,
                r.tipo_dia,
                r.servicio,
                r.codigo_ts,
                r.ruta,
                r.ruta_normalizada,
                r.sentido,
                r.periodo,
                r.expediciones,
                r.buses,

                (
                    round(
                        r.velocidad_real,
                        2
                    )
                    if r.velocidad_real is not None
                    else None
                ),

                (
                    round(
                        r.velocidad_teorica,
                        2
                    )
                    if r.velocidad_teorica is not None
                    else None
                ),

                (
                    round(
                        r.porcentaje_reduccion,
                        2
                    )
                    if r.porcentaje_reduccion is not None
                    else None
                ),

                r.indicador_tiempo_espera,
                r.clasificacion,
                r.estado,

                (
                    "SI"
                    if r.informar
                    else "NO"
                ),

                r.observacion,
                r.archivo_origen,
            ]
        )

    _formatear_cabecera(
        ws
    )

    _ajustar_columnas(
        ws
    )

    # =====================================================
    # HOJA DETALLE PPU
    # =====================================================

    ws_ppu = wb.create_sheet(
        "Detalle PPU"
    )

    encabezados_ppu = [
        "ID PPU",
        "ID Histórico",
        "Fecha",
        "Unidad",
        "Tipo Día",
        "Servicio Usuario",
        "Código TS",
        "Servicio Empresa",
        "Ruta Normalizada",
        "Sentido",
        "Período",
        "Patente",
        "Inicio Servicio",
        "Fin Servicio",
        "Franja Horaria",
        "Velocidad Real",
        "Velocidad Teórica",
        "Reducción %",
        "Indicador",
        "Clasificación",
        "Estado",
        "Archivo Origen",
    ]

    ws_ppu.append(
        encabezados_ppu
    )

    for p in ppus:

        ws_ppu.append(
            [
                p.id,
                p.historico_id,

                (
                    p.fecha_operacional.isoformat()
                    if p.fecha_operacional
                    else None
                ),

                p.unidad,
                p.tipo_dia,
                p.servicio,
                p.codigo_ts,
                p.ruta,
                p.ruta_normalizada,
                p.sentido,
                p.periodo,
                p.patente,

                (
                    p.inicio_servicio.strftime(
                        "%d/%m/%Y %H:%M:%S"
                    )
                    if p.inicio_servicio
                    else None
                ),

                (
                    p.fin_servicio.strftime(
                        "%d/%m/%Y %H:%M:%S"
                    )
                    if p.fin_servicio
                    else None
                ),

                p.franja_horaria,

                (
                    round(
                        p.velocidad_real,
                        2
                    )
                    if p.velocidad_real is not None
                    else None
                ),

                (
                    round(
                        p.velocidad_teorica,
                        2
                    )
                    if p.velocidad_teorica is not None
                    else None
                ),

                (
                    round(
                        p.porcentaje_reduccion,
                        2
                    )
                    if p.porcentaje_reduccion is not None
                    else None
                ),

                p.indicador_tiempo_espera,
                p.clasificacion,
                p.estado,
                p.archivo_origen,
            ]
        )

    _formatear_cabecera(
        ws_ppu
    )

    _ajustar_columnas(
        ws_ppu
    )

    salida = BytesIO()

    wb.save(
        salida
    )

    salida.seek(
        0
    )

    return {
        "archivo": salida,
        "registros": len(
            registros
        ),
        "ppu": len(
            ppus
        ),
    }
